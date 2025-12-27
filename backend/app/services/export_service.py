"""
Export service for generating CSV and Excel files.
"""

import csv
import io
from typing import List, Dict, Any, Optional
from datetime import date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..models.budget import BudgetEntry
from ..models.goal import SavingsGoal
from ..models.category import FinancialCategory


class ExportService:
    """Service class for exporting data to CSV and Excel formats."""

    # Category type display names
    CATEGORY_TYPE_LABELS = {
        "income": "Income",
        "savings": "Savings",
        "expense": "Expense",
    }

    # Expense type display names
    EXPENSE_TYPE_LABELS = {
        "debt_loan": "Debt & Loans",
        "insurance": "Insurance",
        "fixed": "Fixed Expenses",
        "recurring": "Recurring",
        "medical": "Medical",
        "discretionary": "Discretionary",
    }

    # Month names
    MONTH_NAMES = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    @staticmethod
    def _format_decimal(value: Optional[Decimal]) -> str:
        """Format decimal value for export."""
        if value is None:
            return "0.00"
        return f"{value:.2f}"

    @staticmethod
    def _format_date(value: Optional[date]) -> str:
        """Format date value for export."""
        if value is None:
            return ""
        return value.strftime("%Y-%m-%d")

    @classmethod
    def export_budgets_csv(
        cls,
        entries: List[BudgetEntry],
        year: int,
        month: Optional[int] = None
    ) -> str:
        """
        Export budget entries to CSV format.

        Args:
            entries: List of BudgetEntry objects.
            year: Year for the export.
            month: Optional month filter.

        Returns:
            CSV string.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            "Category",
            "Type",
            "Expense Type",
            "Month",
            "Year",
            "Budgeted Amount",
            "Actual Amount",
            "Difference",
            "Notes"
        ])

        # Data rows
        for entry in entries:
            category = entry.category
            category_name = category.name if category else "Unknown"
            category_type = cls.CATEGORY_TYPE_LABELS.get(
                category.category_type.value if category else "",
                "Unknown"
            )
            expense_type = ""
            if category and category.expense_type:
                expense_type = cls.EXPENSE_TYPE_LABELS.get(
                    category.expense_type.value,
                    category.expense_type.value
                )

            budgeted = entry.budgeted_amount or Decimal("0.00")
            actual = entry.actual_amount or Decimal("0.00")
            difference = actual - budgeted

            writer.writerow([
                category_name,
                category_type,
                expense_type,
                cls.MONTH_NAMES[entry.month - 1],
                entry.year,
                cls._format_decimal(budgeted),
                cls._format_decimal(actual),
                cls._format_decimal(difference),
                entry.notes or ""
            ])

        return output.getvalue()

    @classmethod
    def export_budgets_excel(
        cls,
        entries: List[BudgetEntry],
        year: int,
        month: Optional[int] = None
    ) -> bytes:
        """
        Export budget entries to Excel format.

        Args:
            entries: List of BudgetEntry objects.
            year: Year for the export.
            month: Optional month filter.

        Returns:
            Excel file bytes.
        """
        wb = Workbook()
        ws = wb.active

        # Title
        month_str = cls.MONTH_NAMES[month - 1] if month else "All Months"
        ws.title = f"Budget {year}"
        ws.merge_cells("A1:I1")
        ws["A1"] = f"Budget Report - {month_str} {year}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Headers
        headers = [
            "Category", "Type", "Expense Type", "Month", "Year",
            "Budgeted", "Actual", "Difference", "Notes"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, entry in enumerate(entries, 4):
            category = entry.category
            category_name = category.name if category else "Unknown"
            category_type = cls.CATEGORY_TYPE_LABELS.get(
                category.category_type.value if category else "",
                "Unknown"
            )
            expense_type = ""
            if category and category.expense_type:
                expense_type = cls.EXPENSE_TYPE_LABELS.get(
                    category.expense_type.value,
                    category.expense_type.value
                )

            budgeted = float(entry.budgeted_amount or 0)
            actual = float(entry.actual_amount or 0)
            difference = actual - budgeted

            row_data = [
                category_name,
                category_type,
                expense_type,
                cls.MONTH_NAMES[entry.month - 1],
                entry.year,
                budgeted,
                actual,
                difference,
                entry.notes or ""
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col in [6, 7, 8]:  # Currency columns
                    cell.number_format = '"$"#,##0.00'
                    if col == 8:  # Difference column
                        if difference < 0:
                            cell.font = Font(color="FF0000")
                        elif difference > 0:
                            cell.font = Font(color="008000")

        # Auto-adjust column widths
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Summary section
        if entries:
            summary_row = len(entries) + 6

            # Group by type
            income_entries = [e for e in entries if e.category and e.category.category_type.value == "income"]
            savings_entries = [e for e in entries if e.category and e.category.category_type.value == "savings"]
            expense_entries = [e for e in entries if e.category and e.category.category_type.value == "expense"]

            total_income = sum(float(e.actual_amount or e.budgeted_amount or 0) for e in income_entries)
            total_savings = sum(float(e.actual_amount or e.budgeted_amount or 0) for e in savings_entries)
            total_expenses = sum(float(e.actual_amount or e.budgeted_amount or 0) for e in expense_entries)
            surplus = total_income - total_expenses - total_savings

            ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=12)

            summaries = [
                ("Total Income", total_income, "008000"),
                ("Total Savings", total_savings, "0000FF"),
                ("Total Expenses", total_expenses, "FF0000"),
                ("Surplus/Deficit", surplus, "008000" if surplus >= 0 else "FF0000"),
            ]

            for idx, (label, value, color) in enumerate(summaries):
                row = summary_row + idx + 1
                ws.cell(row=row, column=1, value=label)
                cell = ws.cell(row=row, column=2, value=value)
                cell.number_format = '"$"#,##0.00'
                cell.font = Font(color=color, bold=True)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def export_goals_csv(cls, goals: List[SavingsGoal]) -> str:
        """
        Export savings goals to CSV format.

        Args:
            goals: List of SavingsGoal objects.

        Returns:
            CSV string.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            "Goal Name",
            "Target Amount",
            "Current Amount",
            "Remaining Amount",
            "Progress %",
            "Target Date",
            "Status",
            "Created Date"
        ])

        # Data rows
        for goal in goals:
            status = "Completed" if goal.progress_percentage >= 100 else (
                "Active" if goal.is_active else "Inactive"
            )

            writer.writerow([
                goal.name,
                cls._format_decimal(goal.target_amount),
                cls._format_decimal(goal.current_amount),
                cls._format_decimal(goal.remaining_amount),
                f"{goal.progress_percentage:.1f}%",
                cls._format_date(goal.target_date),
                status,
                goal.created_at.strftime("%Y-%m-%d") if goal.created_at else ""
            ])

        return output.getvalue()

    @classmethod
    def export_goals_excel(cls, goals: List[SavingsGoal]) -> bytes:
        """
        Export savings goals to Excel format.

        Args:
            goals: List of SavingsGoal objects.

        Returns:
            Excel file bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Savings Goals"

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = "Savings Goals Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Headers
        headers = [
            "Goal Name", "Target Amount", "Current Amount", "Remaining Amount",
            "Progress %", "Target Date", "Status", "Created Date"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, goal in enumerate(goals, 4):
            status = "Completed" if goal.progress_percentage >= 100 else (
                "Active" if goal.is_active else "Inactive"
            )

            row_data = [
                goal.name,
                float(goal.target_amount or 0),
                float(goal.current_amount or 0),
                float(goal.remaining_amount or 0),
                goal.progress_percentage / 100,  # Store as decimal for percentage format
                goal.target_date,
                status,
                goal.created_at.date() if goal.created_at else None
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col in [2, 3, 4]:  # Currency columns
                    cell.number_format = '"$"#,##0.00'
                elif col == 5:  # Progress percentage
                    cell.number_format = "0.0%"
                    if goal.progress_percentage >= 100:
                        cell.font = Font(color="008000", bold=True)
                elif col in [6, 8]:  # Date columns
                    cell.number_format = "YYYY-MM-DD"
                elif col == 7:  # Status column
                    if value == "Completed":
                        cell.font = Font(color="008000", bold=True)
                    elif value == "Inactive":
                        cell.font = Font(color="808080")

        # Auto-adjust column widths
        column_widths = [20, 15, 15, 15, 12, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Summary section
        if goals:
            summary_row = len(goals) + 6
            active_goals = [g for g in goals if g.is_active]
            completed_goals = [g for g in goals if g.progress_percentage >= 100]

            total_target = sum(float(g.target_amount or 0) for g in active_goals)
            total_current = sum(float(g.current_amount or 0) for g in active_goals)
            overall_progress = (total_current / total_target * 100) if total_target > 0 else 0

            ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=12)

            summaries = [
                ("Total Active Goals", len(active_goals)),
                ("Completed Goals", len(completed_goals)),
                ("Total Target Amount", total_target),
                ("Total Saved Amount", total_current),
                ("Overall Progress", f"{overall_progress:.1f}%"),
            ]

            for idx, (label, value) in enumerate(summaries):
                row = summary_row + idx + 1
                ws.cell(row=row, column=1, value=label)
                cell = ws.cell(row=row, column=2, value=value)
                if isinstance(value, float):
                    cell.number_format = '"$"#,##0.00'

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def export_yearly_summary_csv(
        cls,
        monthly_data: List[Dict[str, Any]],
        year: int
    ) -> str:
        """
        Export yearly summary to CSV format.

        Args:
            monthly_data: List of monthly summary dictionaries.
            year: Year for the export.

        Returns:
            CSV string.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            "Month",
            "Year",
            "Total Income",
            "Total Expenses",
            "Total Savings",
            "Surplus/Deficit"
        ])

        # Data rows
        for data in monthly_data:
            writer.writerow([
                cls.MONTH_NAMES[data["month"] - 1],
                year,
                data.get("total_income", "0.00"),
                data.get("total_expenses", "0.00"),
                data.get("total_savings", "0.00"),
                data.get("surplus_deficit", "0.00")
            ])

        # Totals row
        total_income = sum(float(d.get("total_income", 0)) for d in monthly_data)
        total_expenses = sum(float(d.get("total_expenses", 0)) for d in monthly_data)
        total_savings = sum(float(d.get("total_savings", 0)) for d in monthly_data)
        total_surplus = sum(float(d.get("surplus_deficit", 0)) for d in monthly_data)

        writer.writerow([])
        writer.writerow([
            "TOTAL",
            year,
            f"{total_income:.2f}",
            f"{total_expenses:.2f}",
            f"{total_savings:.2f}",
            f"{total_surplus:.2f}"
        ])

        return output.getvalue()
